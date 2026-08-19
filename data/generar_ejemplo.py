"""
Genera data/ejemplo.xlsx: un extracto ficticio de registro parroquial con
erratas deliberadas, para poder probar la aplicación sin datos reales.

    python data/generar_ejemplo.py

Produce data/ejemplo.xlsx y data/ejemplo.ods con el mismo contenido, para poder
probar la aplicación con los dos formatos que admite.
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# (Any, Llinatge 1, Llinatge 2, Nom, Foli)
#
# Con el vocabulario real del Archivo cargado (core/datos_vocabulario.py, no se
# distribuye con el repositorio), cada errata sembrada ejercita una categoría
# distinta de core.vocabulario.Estado. Sin vocabulario, el programa sigue
# funcionando igual que el prototipo original: todo por heurística de
# frecuencia (ver README, sección "El vocabulario normalizado").
#
#   Benassar -> Bennassar   ninguna de las dos está en el listado real: se
#                           corrige por frecuencia, pero queda marcada "no
#                           verificada" (no es una forma oficial, adendo F).
#   Roselló  -> Rosselló    ORTOGRAFICA: misma palabra, una consonante doble
#                           de menos ("roselo" colapsa igual en las dos).
#   Miquell  -> Miquel      ORTOGRAFICA también, y aparece en un nombre
#                           COMPUESTO: el caso que la versión original perdía.
#   CERDA    -> Cerdà       NORMALIZABLE: mayúsculas y acento, nada más.
#   Rosel    -> ?           AMBIGUA: "Rosell" y "Rossell" son dos apellidos
#                           reales y distintos que colapsan a la misma clave
#                           ortográfica. El programa no elige; lo deja para
#                           revisión humana con las dos opciones a mano.
#
# Trampas que NO deben corregirse (son personas distintas, no erratas):
#   Miquel / Miquela, Antoni / Antònia, Catalina / Catalino
#
# La celda "  'Joan   Miquel " lleva basura de transcripción a propósito:
# la Fase 1 debe dejarla en "Joan Miquel" antes de que empiece la detección.
FILAS = [
    (1789, "Bennassar", "Pons", "  'Joan   Miquel ", "12r"),
    (1789, "Bennassar", "Ferrer", "Antònia", "12v"),
    (1790, "Benassar", "Mas", "Joan Miquell", "13r"),       # 2 erratas en la fila
    (1790, "Bennassar", "Pons", "Miquela", "13v"),          # femenino legítimo
    (1791, "Rosselló", "Vidal", "Miquel", "14r"),
    (1791, "Rosselló", "Cabrer", "Antoni", "14v"),
    (1792, "Roselló", "Vidal", "Joan Miquell", "15r"),      # 2 erratas en la fila
    (1792, "Rosselló", "Mas", "Catalina", "15v"),
    (1793, "Bennassar", "Ferrer", "Miquel Antoni", "16r"),
    (1793, "Rosselló", "Pons", "Joan", "16v"),
    (1794, "Bennassar", "Vidal", "Miquel", "17r"),
    (1794, "Rosselló", "Cabrer", "Catalino", "17v"),        # masculino legítimo
    (1795, "Bennassar", "Mas", "Joan Miquel", "18r"),
    (1795, "Rosselló", "Ferrer", "Antònia", "18v"),
    (1796, "Cerdà", "Ferrer", "Joan", "19r"),
    (1796, "Cerdà", "Vidal", "Antoni", "19v"),
    (1796, "CERDA", "Mas", "Miquel", "20r"),                # NORMALIZABLE
    (1797, "Rosel", "Pons", "Joan", "20v"),                 # AMBIGUA
]

CABECERAS = ("Any", "Llinatge 1", "Llinatge 2", "Nom", "Foli")
TITULO = "Llibre de baptismes — extracte de mostra (dades fictícies)"

# Permite ejecutar el script directamente desde cualquier carpeta.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def crear_xlsx(destino: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Registre"

    # Fila 1: título del libro. Las cabeceras van en la fila 2 (la app usa header=1).
    ws["A1"] = TITULO
    ws["A1"].font = Font(name="Arial", size=12, bold=True)
    ws.merge_cells("A1:E1")

    relleno = PatternFill("solid", fgColor="DDDDDD")
    for col, cabecera in enumerate(CABECERAS, start=1):
        celda = ws.cell(row=2, column=col, value=cabecera)
        celda.font = Font(name="Arial", size=11, bold=True)
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center")

    for fila_idx, fila in enumerate(FILAS, start=3):
        for col_idx, valor in enumerate(fila, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.font = Font(name="Arial", size=11)

    for col, ancho in zip("ABCDE", (8, 16, 16, 20, 8)):
        ws.column_dimensions[col].width = ancho

    wb.save(destino)


def crear_ods(destino: str) -> None:
    """
    Mismo contenido en formato LibreOffice/OpenOffice.

    Se escribe primero la tabla con pandas y después el título de la fila 1
    reutilizando la misma capa de E/S que usa la aplicación.
    """
    import pandas as pd

    from core.workbook_io import abrir_libro

    pd.DataFrame(FILAS, columns=list(CABECERAS)).to_excel(
        destino, index=False, startrow=1, engine="odf"
    )

    libro = abrir_libro(destino)
    libro.escribir_datos(pd.DataFrame([[TITULO]]), primera_fila=1)
    libro.guardar(destino)


def main() -> None:
    carpeta = os.path.dirname(os.path.abspath(__file__))

    ruta_xlsx = os.path.join(carpeta, "ejemplo.xlsx")
    crear_xlsx(ruta_xlsx)
    print(f"Creado: {ruta_xlsx}")

    ruta_ods = os.path.join(carpeta, "ejemplo.ods")
    crear_ods(ruta_ods)
    print(f"Creado: {ruta_ods}")


if __name__ == "__main__":
    main()
