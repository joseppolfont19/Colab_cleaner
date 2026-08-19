"""
Lectura y escritura de libros de cálculo, en XLSX (Excel) y ODS (LibreOffice/OpenOffice).

El resto del programa no necesita saber en qué formato está el archivo: pide un
DataFrame con `leer_tabla()`, y para guardar abre el original con `abrir_libro()`,
le escribe los datos corregidos y lo guarda con otro nombre.

Se trabaja siempre sobre el archivo original en lugar de generar uno nuevo desde
cero, porque así se conservan los formatos, colores, anchos de columna y cualquier
otra hoja que el archivo tuviera. Solo se sobrescriben los valores de las celdas
de datos.
"""

from __future__ import annotations

import datetime as dt
import os
from abc import ABC, abstractmethod

import pandas as pd

# --------------------------------------------------------------------------- #
# Formatos soportados
# --------------------------------------------------------------------------- #

EXTENSION_XLSX = ".xlsx"
EXTENSION_ODS = ".ods"
EXTENSIONES_SOPORTADAS = (EXTENSION_XLSX, EXTENSION_ODS)

# Motores de pandas para cada formato.
_MOTORES = {
    EXTENSION_XLSX: "openpyxl",
    EXTENSION_ODS: "odf",
}

# Para el diálogo de selección de archivo de la interfaz.
FILTROS_DIALOGO = [
    ("Fulls de càlcul", "*.xlsx *.ods"),
    ("Excel", "*.xlsx"),
    ("LibreOffice / OpenOffice", "*.ods"),
]


class FormatoNoSoportado(ValueError):
    """El archivo no es ni .xlsx ni .ods."""


def extension_de(ruta: str) -> str:
    """
    Devuelve la extensión normalizada de un archivo, validando que sea soportada.

    Se compara en minúsculas porque en Windows es frecuente encontrar `.XLSX`.
    """
    ext = os.path.splitext(ruta)[1].lower()
    if ext not in EXTENSIONES_SOPORTADAS:
        soportadas = ", ".join(EXTENSIONES_SOPORTADAS)
        raise FormatoNoSoportado(
            f"Formato no soportado: '{ext or ruta}'. Se admiten: {soportadas}."
        )
    return ext


def leer_tabla(ruta: str, fila_cabecera: int = 1, hoja: str | int = 0) -> pd.DataFrame:
    """
    Carga una hoja como DataFrame, eligiendo el motor según la extensión.

    `hoja` acepta un nombre o un índice (0 = primera/activa, el valor por defecto
    de siempre). Los llamantes que no la usan mantienen el comportamiento actual.
    """
    ext = extension_de(ruta)
    return pd.read_excel(ruta, header=fila_cabecera, sheet_name=hoja, engine=_MOTORES[ext])


def ruta_de_salida(ruta_original: str, sufijo: str = "_corregit") -> str:
    """`registre.ods` -> `registre_corregit.ods`, conservando la extensión."""
    base, ext = os.path.splitext(ruta_original)
    return f"{base}{sufijo}{ext}"


# --------------------------------------------------------------------------- #
# Interfaz común
# --------------------------------------------------------------------------- #

class LibroOriginal(ABC):
    """
    Un libro de cálculo abierto, listo para recibir valores y guardarse aparte.

    Las subclases encapsulan las diferencias entre openpyxl (XLSX) y odfpy (ODS).
    """

    @abstractmethod
    def escribir_datos(self, df: pd.DataFrame, primera_fila: int) -> None:
        """
        Vuelca el DataFrame en la hoja, empezando por `primera_fila` (base 1).

        Solo se tocan las celdas de datos: las cabeceras y cualquier fila
        anterior quedan como estaban.
        """

    @abstractmethod
    def guardar(self, ruta: str) -> None:
        """Escribe el libro en `ruta`."""


def abrir_libro(ruta: str) -> LibroOriginal:
    """Abre el archivo con el motor que corresponda a su extensión."""
    ext = extension_de(ruta)
    if ext == EXTENSION_XLSX:
        return LibroXlsx(ruta)
    return LibroOds(ruta)


def _normalizar_valor(valor):
    """
    Convierte un valor de pandas en algo que una hoja de cálculo pueda almacenar.

    pandas usa NaN para las celdas vacías; escrito tal cual acabaría como el
    texto "nan" en el archivo final.
    """
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        # pd.isna revienta con algunos tipos exóticos (listas, por ejemplo).
        pass
    return valor


# --------------------------------------------------------------------------- #
# Implementación XLSX (openpyxl)
# --------------------------------------------------------------------------- #

class LibroXlsx(LibroOriginal):
    def __init__(self, ruta: str):
        from openpyxl import load_workbook

        self._wb = load_workbook(ruta)
        self._ws = self._wb.active

    def escribir_datos(self, df: pd.DataFrame, primera_fila: int) -> None:
        for fila_idx, fila in enumerate(df.itertuples(index=False), start=primera_fila):
            for col_idx, valor in enumerate(fila, start=1):
                self._ws.cell(row=fila_idx, column=col_idx).value = _normalizar_valor(valor)

    def guardar(self, ruta: str) -> None:
        self._wb.save(ruta)


# --------------------------------------------------------------------------- #
# Implementación ODS (odfpy)
# --------------------------------------------------------------------------- #

# Espacio de nombres de los atributos de tabla en OpenDocument.
_NS_TABLE = "urn:oasis:names:tc:opendocument:xmlns:table:1.0"
_NS_OFFICE = "urn:oasis:names:tc:opendocument:xmlns:office:1.0"

_ATTR_REPETIR_COLS = (_NS_TABLE, "number-columns-repeated")
_ATTR_REPETIR_FILAS = (_NS_TABLE, "number-rows-repeated")
_ATTR_TIPO_VALOR = (_NS_OFFICE, "value-type")
_ATTR_VALOR = (_NS_OFFICE, "value")
_ATTR_VALOR_FECHA = (_NS_OFFICE, "date-value")
_ATTR_VALOR_BOOL = (_NS_OFFICE, "boolean-value")

_ATTRS_DE_VALOR = (_ATTR_TIPO_VALOR, _ATTR_VALOR, _ATTR_VALOR_FECHA, _ATTR_VALOR_BOOL)

_QNAME_FILA = (_NS_TABLE, "table-row")
_QNAMES_CELDA = ((_NS_TABLE, "table-cell"), (_NS_TABLE, "covered-table-cell"))


class LibroOds(LibroOriginal):
    """
    Escritura sobre un ODS existente.

    La dificultad del formato ODS es que comprime las celdas y filas repetidas:
    en vez de guardar diez celdas vacías iguales, guarda una con el atributo
    `number-columns-repeated="10"`. Para escribir en la celda 3 de ese bloque hay
    que partirlo antes en tres trozos. De eso se ocupan `_fila_en` y `_celda_en`.
    """

    def __init__(self, ruta: str):
        from odf.opendocument import load
        from odf.table import Table

        self._doc = load(ruta)
        tablas = self._doc.spreadsheet.getElementsByType(Table)
        if not tablas:
            raise ValueError("El archivo ODS no contiene ninguna hoja.")
        self._tabla = tablas[0]

    # -- utilidades de acceso ------------------------------------------------ #

    @staticmethod
    def _repeticiones(nodo, atributo) -> int:
        valor = nodo.attributes.get(atributo)
        try:
            return max(1, int(valor))
        except (TypeError, ValueError):
            return 1

    @staticmethod
    def _texto_de(nodo) -> str:
        return str(nodo) if nodo.childNodes else ""

    @classmethod
    def _clonar(cls, plantilla, repeticiones: int, constructor, atributo_repetir):
        """Crea un nodo con los mismos atributos que la plantilla, salvo la repetición."""
        nuevo = constructor()
        for clave, valor in plantilla.attributes.items():
            if clave != atributo_repetir:
                nuevo.attributes[clave] = valor
        if repeticiones > 1:
            nuevo.attributes[atributo_repetir] = str(repeticiones)
        return nuevo

    @staticmethod
    def _insertar_y_registrar(contenedor, nuevo, referencia) -> None:
        """
        `contenedor.insertBefore(nuevo, referencia)`, seguido de registrar
        `nuevo` en el documento a mano.

        Bug real de odfpy (hay un comentario "FIXME: update
        ownerDocument.element_dict or find other solution" en su propio
        `Element.removeChild`): `insertBefore()` no hace lo mismo que
        `addElement()` -- no asigna `ownerDocument` al nodo insertado ni lo
        da de alta en el índice interno del documento (`element_dict`).
        Mientras ese nodo no se vuelva a tocar, no pasa nada. Pero si una
        fila tiene un bloque de celdas repetidas que hay que partir en MÁS
        de un punto (dos columnas distintas caen dentro del mismo bloque
        original), el segundo `_partir()` cae sobre un trozo creado por el
        primero -- un nodo insertado con `insertBefore()`, nunca registrado
        -- y al intentar quitarlo con `removeChild()`, odfpy intenta
        borrarlo de una lista donde nunca estuvo:
        `ValueError: list.remove(x): x not in list`.

        Se repite aquí, a mano, exactamente lo que hace `addElement()`
        (`Element._setOwnerDoc()` + `Document.rebuild_caches()`) para que
        cualquier trozo, se cree cuando se cree, quede tan "registrado" como
        si hubiera estado en el documento desde el principio.
        """
        contenedor.insertBefore(nuevo, referencia)
        documento = contenedor.ownerDocument
        if documento is not None:
            contenedor._setOwnerDoc(nuevo)
            documento.rebuild_caches(nuevo)

    def _partir(self, contenedor, nodos, indice, constructor, atributo_repetir):
        """
        Devuelve el nodo que ocupa `indice`, partiendo los bloques repetidos si hace falta.

        Si el índice cae más allá del último nodo, se añaden nodos vacíos hasta llegar.
        """
        posicion = 0
        for nodo in nodos:
            repeticiones = self._repeticiones(nodo, atributo_repetir)
            if posicion <= indice < posicion + repeticiones:
                if repeticiones == 1:
                    return nodo

                desplazamiento = indice - posicion
                anteriores = desplazamiento
                posteriores = repeticiones - desplazamiento - 1

                objetivo = self._clonar(nodo, 1, constructor, atributo_repetir)
                if anteriores:
                    self._insertar_y_registrar(
                        contenedor, self._clonar(nodo, anteriores, constructor, atributo_repetir), nodo
                    )
                self._insertar_y_registrar(contenedor, objetivo, nodo)
                if posteriores:
                    self._insertar_y_registrar(
                        contenedor, self._clonar(nodo, posteriores, constructor, atributo_repetir), nodo
                    )
                contenedor.removeChild(nodo)
                return objetivo
            posicion += repeticiones

        # El índice queda fuera: rellenar con nodos vacíos hasta alcanzarlo.
        while posicion <= indice:
            nuevo = constructor()
            contenedor.addElement(nuevo)
            posicion += 1
        return nuevo

    def _filas(self):
        # odfpy expone TableRow como función constructora, no como clase, así que
        # los nodos se identifican por su nombre cualificado y no con isinstance.
        return [n for n in self._tabla.childNodes if n.qname == _QNAME_FILA]

    def _fila_en(self, indice: int):
        from odf.table import TableRow

        return self._partir(
            self._tabla, self._filas(), indice, TableRow, _ATTR_REPETIR_FILAS
        )

    def _celda_en(self, fila, indice: int):
        from odf.table import TableCell

        # Las celdas cubiertas por una combinación también ocupan posición.
        celdas = [n for n in fila.childNodes if n.qname in _QNAMES_CELDA]
        return self._partir(fila, celdas, indice, TableCell, _ATTR_REPETIR_COLS)

    @staticmethod
    def _escribir_celda(celda, valor) -> None:
        """Sustituye el contenido de una celda conservando su estilo."""
        from odf.text import P

        for hijo in list(celda.childNodes):
            celda.removeChild(hijo)
        for atributo in _ATTRS_DE_VALOR:
            celda.attributes.pop(atributo, None)

        if valor is None:
            return

        if isinstance(valor, bool):
            celda.attributes[_ATTR_TIPO_VALOR] = "boolean"
            celda.attributes[_ATTR_VALOR_BOOL] = "true" if valor else "false"
            texto = "TRUE" if valor else "FALSE"
        elif isinstance(valor, (dt.datetime, dt.date, pd.Timestamp)):
            celda.attributes[_ATTR_TIPO_VALOR] = "date"
            celda.attributes[_ATTR_VALOR_FECHA] = valor.isoformat()
            texto = valor.isoformat()
        else:
            try:
                numero = float(valor)
            except (TypeError, ValueError):
                celda.attributes[_ATTR_TIPO_VALOR] = "string"
                texto = str(valor)
            else:
                celda.attributes[_ATTR_TIPO_VALOR] = "float"
                celda.attributes[_ATTR_VALOR] = repr(numero)
                # Los enteros se muestran sin el ".0" que arrastra float().
                texto = str(int(numero)) if numero.is_integer() else str(numero)

        celda.addElement(P(text=texto))

    # -- interfaz pública ---------------------------------------------------- #

    def escribir_datos(self, df: pd.DataFrame, primera_fila: int) -> None:
        for desplazamiento, datos in enumerate(df.itertuples(index=False)):
            fila = self._fila_en(primera_fila - 1 + desplazamiento)
            for col_idx, valor in enumerate(datos):
                celda = self._celda_en(fila, col_idx)
                self._escribir_celda(celda, _normalizar_valor(valor))

    def guardar(self, ruta: str) -> None:
        self._doc.save(ruta)
